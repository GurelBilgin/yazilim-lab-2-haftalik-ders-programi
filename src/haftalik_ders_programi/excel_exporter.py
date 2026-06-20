
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .scheduler import DAYS, HOURS, SINIFLAR, Schedule


def export_schedule(schedule: Schedule, output_path: str | Path) -> Path:
    """Haftalık ders programını Excel dosyasına yazar."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Haftalık Program"

    ws.append(["Bölüm", "", schedule.baslik, "", "", ""])
    ws.append(["Gün", "Saat", "1. Sınıf", "2. Sınıf", "3. Sınıf", "4. Sınıf"])

    for gun in DAYS:
        for saat_index, saat in enumerate(HOURS):
            ws.append([gun if saat_index == 0 else "", saat, "", "", "", ""])

    for row_idx in range(3, ws.max_row + 1):
        gun = ws.cell(row=row_idx, column=1).value
        if not gun:
            for previous in range(row_idx - 1, 2, -1):
                gun = ws.cell(row=previous, column=1).value
                if gun:
                    break

        saat = ws.cell(row=row_idx, column=2).value
        for col_idx, sinif in enumerate(SINIFLAR, start=3):
            ws.cell(row=row_idx, column=col_idx, value=schedule.get_cell(gun, saat, sinif).metin)

    _style_sheet(ws)
    wb.save(output_path)
    return output_path


def _style_sheet(ws) -> None:
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    sub_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    day_fill = PatternFill(start_color="F4F7FB", end_color="F4F7FB", fill_type="solid")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("C1:F1")
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.border = border
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["C1"].font = Font(bold=True, color="FFFFFF", size=14)

    for col in range(1, 7):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.fill = sub_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 18,
        "B": 18,
        "C": 38,
        "D": 38,
        "E": 38,
        "F": 38,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 45 if row > 2 else 28
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(3, ws.max_row + 1, len(HOURS)):
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = day_fill

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:F{ws.max_row}"
