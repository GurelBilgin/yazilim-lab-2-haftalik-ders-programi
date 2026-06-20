
from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook

from haftalik_ders_programi.excel_exporter import export_schedule
from haftalik_ders_programi.scheduler import ScheduleGenerator
from haftalik_ders_programi.seed_loader import load_program_dersleri


SQL_DIR = Path(__file__).resolve().parents[1] / "sql"


class ExcelExporterTestleri(unittest.TestCase):
    def test_excel_ciktisi_olusturur(self):
        dersler = load_program_dersleri(SQL_DIR)
        generator = ScheduleGenerator()
        schedule = generator.generate_for_department(dersler, "BLM", "Bilgisayar Mühendisliği")

        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "program.xlsx"
            export_schedule(schedule, output_path)

            self.assertTrue(output_path.exists())

            wb = load_workbook(output_path)
            ws = wb.active
            self.assertEqual(ws.title, "Haftalık Program")
            self.assertEqual(ws["A2"].value, "Gün")
            self.assertEqual(ws["C2"].value, "1. Sınıf")
            self.assertGreater(ws.max_row, 10)


if __name__ == "__main__":
    unittest.main()
